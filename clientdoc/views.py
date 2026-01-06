# clientdoc/views.py

from django.template.loader import render_to_string
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction
from django.core.paginator import Paginator
from django.conf import settings
from django.urls import reverse
from .models import SalesInvoice, InvoiceItem, Item, StoreLocation, DeliveryChallan, TransportCharges, ConfirmationDocument, PackedImage, OurCompanyProfile, ActivityLog, Buyer, BulkInvoiceUpload, ItemCategory
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from .forms import InvoiceForm, DeliveryChallanForm, TransportChargesForm, ConfirmationDocumentForm, PackedImageFormSet, ItemForm, StoreLocationForm, BuyerForm, InvoiceItemFormSet
import json
from .pdf_generator import generate_invoice_pdf, generate_dc_pdf, generate_transport_pdf
import logging
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as PlatypusImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from PyPDF2 import PdfMerger, PdfReader
import os

logger = logging.getLogger(__name__)

# --- PDF GENERATION HELPERS ---

def generate_packed_images_pdf(confirmation):
    """Generates a PDF page for packed images."""
    images = confirmation.packedimage_set.all()
    if not images.exists():
        return None 

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 50
    img_width = width - (2 * margin) 
    img_height = 250 
    spacing = 20
    y = height - margin
    
    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, y, "Packed Goods Images")
    y -= 40
    
    for image_obj in images:
        if y < margin + img_height + spacing:
            c.showPage()
            y = height - margin - 20 
            
        try:
            img_path = image_obj.image.path
            img = ImageReader(img_path) 
            
            aspect = img.getSize()[1] / img.getSize()[0]
            current_img_height = img_width * aspect
            
            if current_img_height > img_height:
                current_img_height = img_height

            c.drawImage(img, margin, y - current_img_height, width=img_width, height=current_img_height)
            
            c.setFont("Helvetica", 10)
            notes_y = y - current_img_height - 10
            c.drawString(margin, notes_y, f"Notes: {image_obj.notes or 'N/A'}")

            y -= (current_img_height + spacing + 20) 

        except Exception as e:
            logger.error(f"Error drawing image {image_obj.id} to PDF: {e}")
            c.setFont("Helvetica-Bold", 12)
            c.drawString(margin, y, f"Error loading image {image_obj.id}: {e}")
            y -= 30
            
    c.save()
    buffer.seek(0)
    return buffer


# --- 1. DASHBOARD & LIST VIEWS (FIX 3: Corrected List Views) ---

def dashboard(request):
    """Shows system overview and recent activity (recent invoices)."""
    invoices = SalesInvoice.objects.all().select_related('location').order_by('-date')[:10]
    total_invoices = SalesInvoice.objects.count()
    total_finalized = SalesInvoice.objects.filter(status='FIN').count()
    
    context = {
        'invoices': invoices,
        'total_invoices': total_invoices,
        'total_finalized': total_finalized
    }
    
    recent_logs = ActivityLog.objects.order_by('-timestamp')[:10]
    
    context['recent_logs'] = recent_logs
    return render(request, 'clientdoc/dashboard.html', context)

def item_detail(request, item_id):
    """Detail view for a single item."""
    item = get_object_or_404(Item, id=item_id)
    return render(request, 'clientdoc/item_detail.html', {'item': item})

def log_activity(action, details=""):
    ActivityLog.objects.create(action=action, details=details)

def get_filtered_queryset(model_class, request, search_fields):
    """Helper to filter and sort querysets."""
    queryset = model_class.objects.all().select_related('invoice') if model_class != SalesInvoice and hasattr(model_class, 'invoice') else model_class.objects.all()
    if model_class == SalesInvoice:
        queryset = queryset.select_related('location')
    elif hasattr(model_class, 'invoice'):
         queryset = queryset.filter(invoice__is_deleted=False)
        
    # Search
    query = request.GET.get('q')
    if query:
        from django.db.models import Q
        q_objects = Q()
        for field in search_fields:
            q_objects |= Q(**{field + '__icontains': query})
        queryset = queryset.filter(q_objects)
    
    # Sort
    # Sort
    sort_by = request.GET.get('sort')
    
    # Determine default sort if not provided
    if not sort_by:
        if hasattr(model_class, 'date'):
            sort_by = '-date'
        elif hasattr(model_class, 'created_at'):
            sort_by = '-created_at'
        else:
            sort_by = '-id'

    if sort_by == 'az': 
        if model_class == SalesInvoice:
            sort_by = 'tally_invoice_number'
        elif hasattr(model_class, 'name'):
            sort_by = 'name'
        else:
            sort_by = 'invoice__tally_invoice_number'

    if sort_by == 'za': 
        if model_class == SalesInvoice:
            sort_by = '-tally_invoice_number'
        elif hasattr(model_class, 'name'):
            sort_by = '-name'
        else:
            sort_by = '-invoice__tally_invoice_number'
            
    # Better date sorting using created_at if available and date is not
    if sort_by in ['date', '-date'] and not hasattr(model_class, 'date') and hasattr(model_class, 'created_at'):
        sort_by = sort_by.replace('date', 'created_at')
        
    # Safety check: If trying to sort by date/created_at but model lacks it
    if 'date' in sort_by and not hasattr(model_class, 'date'):
        sort_by = '-id'
    if 'created_at' in sort_by and not hasattr(model_class, 'created_at'):
        sort_by = '-id'
    
    allowed_sorts = [
         'date', '-date', 
         'created_at', '-created_at', 
         'id', '-id', 
         'total', '-total', 
         'status', '-status', 
         'tally_invoice_number', '-tally_invoice_number', 
         'app_invoice_number', '-app_invoice_number',
         'invoice__tally_invoice_number', '-invoice__tally_invoice_number', 
         'invoice__app_invoice_number', '-invoice__app_invoice_number',
         'invoice__date', '-invoice__date',
         'name', '-name'
    ]
                     
    if sort_by in allowed_sorts:
        queryset = queryset.order_by(sort_by)
    else:
        # Default sorts
        if hasattr(model_class, 'date'):
            queryset = queryset.order_by('-date')
        elif hasattr(model_class, 'invoice'):
             queryset = queryset.order_by('-invoice__date')
        else:
             queryset = queryset.order_by('-id')
        
    return queryset

def trash_list(request):
    """View to show deleted items."""
    invoices = SalesInvoice.objects.trash().all()
    locations = StoreLocation.objects.trash().all()
    items = Item.objects.trash().all()
    
    return render(request, 'clientdoc/trash_list.html', {
        'invoices': invoices,
        'locations': locations,
        'items': items,
        'title': 'Trash Bin'
    })

def restore_object(request, model_name, pk):
    """Restores a soft-deleted object."""
    model_map = {
        'invoice': SalesInvoice,
        'location': StoreLocation,
        'item': Item,
        'dc': DeliveryChallan,
        'transport': TransportCharges,
        'confirmation': ConfirmationDocument,
        'buyer': Buyer
    }
    model = model_map.get(model_name)
    if not model:
        messages.error(request, 'Invalid item type.')
        return redirect('clientdoc:trash_list')
        
    obj = get_object_or_404(model.objects.trash(), pk=pk)
    obj.restore()
    log_activity("Restore", f"Restored {model_name} #{pk}")
    messages.success(request, f'{model_name.title()} restored successfully.')
    return redirect('clientdoc:trash_list')

def hard_delete_object(request, model_name, pk):
    """Permanently deletes an object."""
    model_map = {
        'invoice': SalesInvoice,
        'location': StoreLocation,
        'item': Item,
        'dc': DeliveryChallan,
        'transport': TransportCharges,
        'confirmation': ConfirmationDocument,
        'buyer': Buyer
    }
    model = model_map.get(model_name)
    if not model:
        messages.error(request, 'Invalid item type.')
        return redirect('clientdoc:trash_list')
        
    obj = get_object_or_404(model.objects.trash(), pk=pk)
    obj.hard_delete()
    log_activity("Permanent Delete", f"Permanently deleted {model_name} #{pk}")
    messages.warning(request, f'{model_name.title()} permanently deleted.')
    return redirect('clientdoc:trash_list')

def delete_object(request, model_name, pk):
    """Soft deletes an object from list view."""
    model_map = {
        'invoice': SalesInvoice,
        'location': StoreLocation,
        'item': Item,
        'dc': DeliveryChallan,
        'transport': TransportCharges,
        'confirmation': ConfirmationDocument,
        'buyer': Buyer
    }
    model = model_map.get(model_name)
    if not model:
        messages.error(request, 'Invalid item type.')
        return redirect('clientdoc:dashboard')

    obj = get_object_or_404(model, pk=pk)
    obj.delete() # Soft delete
    log_activity("Delete", f"Moved {model_name} #{pk} to trash")
    messages.success(request, f'{model_name.title()} moved to trash.')
    return redirect(request.META.get('HTTP_REFERER', 'clientdoc:dashboard'))

def invoice_list(request):
    search_fields = ['tally_invoice_number', 'app_invoice_number', 'location__name', 'date']
    invoices = get_filtered_queryset(SalesInvoice, request, search_fields)
    
    paginator = Paginator(invoices, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/invoice_list.html', {
        'page_obj': page_obj, 
        'title': 'Sales Invoice List',
        'list_type': 'inv'
    })

def dc_list(request):
    search_fields = ['invoice__tally_invoice_number', 'invoice__app_invoice_number', 'invoice__location__name', 'date']
    challans = get_filtered_queryset(DeliveryChallan, request, search_fields)
    
    paginator = Paginator(challans, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/dc_list.html', {
        'page_obj': page_obj, 
        'title': 'Delivery Challan List',
        'list_type': 'dc'
    })
    
def transport_list(request):
    search_fields = ['invoice__tally_invoice_number', 'invoice__app_invoice_number', 'invoice__location__name', 'date', 'description']
    charges = get_filtered_queryset(TransportCharges, request, search_fields)
    
    paginator = Paginator(charges, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/transport_list.html', {
        'page_obj': page_obj, 
        'title': 'Transport Charges List',
        'list_type': 'trp'
    })

def confirmation_list(request):
    search_fields = ['invoice__tally_invoice_number', 'invoice__app_invoice_number', 'invoice__location__name', 'date']
    docs = get_filtered_queryset(ConfirmationDocument, request, search_fields)
    
    paginator = Paginator(docs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/confirmation_list.html', {
        'page_obj': page_obj, 
        'title': 'Confirmation Document List',
        'list_type': 'cnf'
    })

# --- Utility Functions (Ensure create_item exists) ---

# --- ITEM VIEWS ---

def item_list(request):
    search_fields = ['name', 'description']
    items = get_filtered_queryset(Item, request, search_fields)
    
    paginator = Paginator(items, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/item_list.html', {
        'page_obj': page_obj, 
        'title': 'Item List',
        'list_type': 'item'
    })

def edit_item(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            log_activity("Edit Item", f"Updated Item {item.name}")
            messages.success(request, 'Item updated successfully.')
            return redirect('clientdoc:item_list')
    else:
        form = ItemForm(instance=item)
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Edit Item'})

def create_item(request):
    if request.method == 'POST':
        form = ItemForm(request.POST) 
        if form.is_valid():
            form.save()
            log_activity("Create Item", f"Created Item {form.instance.name}")
            messages.success(request, 'Item created successfully.')
            return redirect('clientdoc:dashboard')
    else:
        form = ItemForm()
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Create Item'})

def create_location(request):
    if request.method == 'POST':
        form = StoreLocationForm(request.POST) 
        if form.is_valid():
            form.save()
            log_activity("Create Location", f"Created Location {form.instance.name}")
            messages.success(request, 'Location created successfully.')
            return redirect('clientdoc:dashboard')
    else:
        form = StoreLocationForm()
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Create Store Location'})

def store_location_list(request):
    search_fields = ['name', 'address', 'city', 'gstin', 'site_code']
    locations = get_filtered_queryset(StoreLocation, request, search_fields)
    
    paginator = Paginator(locations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/store_location_list.html', {
        'page_obj': page_obj, 
        'title': 'Store Client Locations',
        'list_type': 'location'
    })

def edit_location(request, pk):
    location = get_object_or_404(StoreLocation, pk=pk)
    if request.method == 'POST':
        form = StoreLocationForm(request.POST, instance=location)
        if form.is_valid():
            form.save()
            log_activity("Edit Location", f"Updated Location {location.name}")
            messages.success(request, 'Location updated successfully.')
            return redirect('clientdoc:store_location_list')
    else:
        form = StoreLocationForm(instance=location)
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Edit Store Location'})

def store_location_detail(request, pk):
    location = get_object_or_404(StoreLocation, pk=pk)
    return render(request, 'clientdoc/store_location_detail.html', {'location': location})

# --- 2. WORKFLOW STEP 1: CREATE INVOICE ITEMS ---

def create_invoice(request):
    """Handles creation of SalesInvoice and multiple InvoiceItem records using FormSets."""
    
    items = Item.objects.all()
    locations = StoreLocation.objects.all()
    buyers = Buyer.objects.all()
    
    # Use the imported FormSet, or create a factory if specific config needed
    # formset = InvoiceItemFormSet(queryset=InvoiceItem.objects.none()) # If usage of imported one

    if request.method == 'POST':
        location_id = request.POST.get('location')
        buyer_id = request.POST.get('buyer')
        tally_invoice_number = request.POST.get('tally_invoice_number')
        date = request.POST.get('date')

        if not location_id:
            messages.error(request, 'Please select a client location.')
            return redirect('clientdoc:create_invoice')

        try:
            with transaction.atomic():
                location = get_object_or_404(StoreLocation, id=location_id)
                buyer = None
                if buyer_id:
                    buyer = get_object_or_404(Buyer, id=buyer_id)
                
                invoice = SalesInvoice.objects.create(location=location, buyer=buyer, status='DRF') 
                if tally_invoice_number: invoice.tally_invoice_number = tally_invoice_number
                if date: invoice.date = date
                invoice.save()
                
                # Bind formset to new invoice
                formset = InvoiceItemFormSet(request.POST, instance=invoice, prefix='invoiceitem_set')
                
                print(f"DEBUG: TOTAL_FORMS: {request.POST.get('invoiceitem_set-TOTAL_FORMS')}")
                if not formset.is_valid():
                    print(f"DEBUG: Formset Errors: {formset.errors}")
                    print(f"DEBUG: NonForm Errors: {formset.non_form_errors()}")

                print(f"DEBUG: TOTAL_FORMS: {request.POST.get('invoiceitem_set-TOTAL_FORMS')}")
                if formset.is_valid():
                    # Standard save handles foreign keys because instance=invoice is passed
                    formset.save()
                    
                    invoice.refresh_from_db()
                    invoice.calculate_total()
                    
                    log_activity("Create Invoice", f"Created Invoice {invoice.id} with {invoice.invoiceitem_set.count()} items")
                    messages.success(request, f'Invoice #{invoice.id} created successfully!')
                    return redirect('clientdoc:edit_invoice', invoice_id=invoice.id)
                else:
                    # Rollback if items are invalid
                    transaction.set_rollback(True)
                    # Show errors
                    if formset.non_form_errors():
                        messages.error(request, f"Formset Error: {formset.non_form_errors()}")
                    for form in formset:
                        for field, errors in form.errors.items():
                             for error in errors:
                                 messages.error(request, f"Item Error ({field}): {error}")
                    messages.error(request, 'Failed to create invoice. Please check item details.')

        except Exception as e:
            logger.error(f"Invoice Create Error: {e}")
            messages.error(request, f"Error creating invoice: {str(e)}")

    else:
        # GET request - Initialize empty formset so we have management form
        # We pass instance=None or a dummy unsaved instance? 
        # inlineformset factory expects instance. SalesInvoice() is fine.
        formset = InvoiceItemFormSet(instance=SalesInvoice(), prefix='invoiceitem_set')
        formset.extra = 0

    return render(request, 'clientdoc/invoice_form.html', {
        'locations': locations,
        'buyers': buyers,
        'items': items,
        'formset': formset,
        'title': 'Create Sales Invoice'
    })


# --- 3. WORKFLOW STEP 2: EDIT INVOICE (TALLY DETAILS) ---
def edit_invoice(request, invoice_id):
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    items = Item.objects.all() 
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST, instance=invoice) 
        formset = InvoiceItemFormSet(request.POST, instance=invoice, prefix='invoiceitem_set')
        
        if form.is_valid() and formset.is_valid():
            form.save()
            
            instances = formset.save(commit=False)
            for instance in instances:
                if instance.item_id:
                    instance.invoice = invoice
                    instance.save()
            
            for obj in formset.deleted_objects:
                obj.delete()
                
            invoice.calculate_total()
            log_activity("Edit Invoice", f"Updated Invoice {invoice.tally_invoice_number or invoice.id} details")
            messages.success(request, f'Invoice details updated.')
            
            if request.POST.get('action') == 'save_continue':
                return redirect('clientdoc:edit_dc', invoice_id=invoice.id)
            if request.POST.get('action') == 'save_list':
                return redirect('clientdoc:invoice_list')
            
            return redirect('clientdoc:edit_invoice', invoice_id=invoice.id) 
        else:
             if not form.is_valid():
                 messages.error(request, f"Header Errors: {form.errors}")
             if not formset.is_valid():
                 messages.error(request, f"Item Errors: {formset.errors}")
    else:
        form = InvoiceForm(instance=invoice)
        formset = InvoiceItemFormSet(instance=invoice, prefix='invoiceitem_set')
    
    invoice.refresh_from_db()
    next_url = reverse('clientdoc:edit_dc', kwargs={'invoice_id': invoice.id})

    return render(request, 'clientdoc/edit_tally_details.html', {
        'form': form,
        'formset': formset,
        'items': items,
        'invoice': invoice, # Fixed
        'title': f'Edit Tally Details for Invoice #{invoice.id}',
        'next_url': next_url, 
        'current_step': 1,
        'progress_percentage': 25,
    })


# --- 4. WORKFLOW STEP 3: EDIT DELIVERY CHALLAN (DC) ---
def edit_dc(request, invoice_id):
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    dc, created = DeliveryChallan.objects.get_or_create(invoice=invoice)

    if request.method == 'POST':
        form = DeliveryChallanForm(request.POST, instance=dc)
        if form.is_valid():
            form.save()
            log_activity("Edit DC", f"Updated DC for Invoice {invoice.id}")
            
            if invoice.status == 'DRF':
                invoice.status = 'DC'
                invoice.save()
                
            # FIX 2: Redirect to the DC List after edit
            messages.success(request, 'Delivery Challan updated.')
            
            if request.POST.get('action') == 'save_continue':
                return redirect('clientdoc:edit_transport', invoice_id=invoice.id)
            if request.POST.get('action') == 'save_list':
                return redirect('clientdoc:dc_list')
                
            return redirect('clientdoc:edit_dc', invoice_id=invoice.id) 
    else:
        form = DeliveryChallanForm(instance=dc)
    
    next_url = reverse('clientdoc:edit_transport', kwargs={'invoice_id': invoice.id})
    prev_url = reverse('clientdoc:edit_invoice', kwargs={'invoice_id': invoice.id})

    return render(request, 'clientdoc/form.html', {
        'form': form,
        'title': f'Delivery Challan - Invoice {invoice.tally_invoice_number or invoice.id}',
        'next_url': next_url,
        'prev_url': prev_url, 
        'current_step': 2,
        'progress_percentage': 50,
    })


# --- 5. WORKFLOW STEP 4: EDIT TRANSPORT CHARGES ---
def edit_transport(request, invoice_id):
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    transport, created = TransportCharges.objects.get_or_create(invoice=invoice)
        
    if invoice.status not in ['DC', 'TRP', 'FIN']:
        messages.error(request, 'You must complete the Delivery Challan first.')
        return redirect('clientdoc:dashboard')
        
    if request.method == 'POST':
        form = TransportChargesForm(request.POST, instance=transport)
        if form.is_valid():
            form.save()
            log_activity("Edit Transport", f"Updated Transport Charges for Invoice {invoice.id}")
            
            if invoice.status == 'DC':
                invoice.status = 'TRP'
                invoice.save()
                
            # FIX 2: Redirect to the Transport Charges List after edit
            messages.success(request, 'Transport charges updated.')
            
            if request.POST.get('action') == 'save_continue':
                return redirect('clientdoc:create_confirmation', invoice_id=invoice.id)
            if request.POST.get('action') == 'save_list':
                return redirect('clientdoc:transport_list')
                
            return redirect('clientdoc:edit_transport', invoice_id=invoice.id) 
    else:
        form = TransportChargesForm(instance=transport)
    
    next_url = reverse('clientdoc:create_confirmation', kwargs={'invoice_id': invoice.id})
    prev_url = reverse('clientdoc:edit_dc', kwargs={'invoice_id': invoice.id})

    return render(request, 'clientdoc/form.html', {
        'form': form,
        'title': f'Transport Charges - Invoice {invoice.tally_invoice_number or invoice.id}',
        'next_url': next_url,
        'prev_url': prev_url, 
        'current_step': 3,
        'progress_percentage': 75,
    })


# --- 6. WORKFLOW STEP 5: CONFIRMATION & PDF GENERATION (FIX 1: Robust Merging) ---

def create_confirmation(request, invoice_id):
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    confirmation, created = ConfirmationDocument.objects.get_or_create(invoice=invoice)
    company_profile = OurCompanyProfile.objects.first() 
    
    if invoice.status not in ['TRP', 'FIN']:
        messages.error(request, 'Cannot access Confirmation Document yet. Please log Transport Charges first.')
        return redirect('clientdoc:dashboard')
    
    # File deletion logic (kept short for brevity)
    if request.method == 'POST':
        if 'delete_po' in request.POST and confirmation.po_file:
            confirmation.po_file.delete(save=False)
            confirmation.po_file = None
            confirmation.save()
            messages.success(request, 'Purchase Order file removed.')
            return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)

        if 'delete_email' in request.POST and confirmation.approval_email_file:
            confirmation.approval_email_file.delete(save=False)
            confirmation.approval_email_file = None
            confirmation.save()
            messages.success(request, 'Approval Email file removed.')
            return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)
    
    has_po = bool(confirmation.po_file)
    has_email = bool(confirmation.approval_email_file)

    if request.method == 'POST':
        form = ConfirmationDocumentForm(request.POST, request.FILES, instance=confirmation)
        image_formset = PackedImageFormSet(request.POST, request.FILES, instance=confirmation)
        
        if form.is_valid() and image_formset.is_valid():
            confirmation = form.save()
            image_formset.save()
            
            if 'save_notes' in request.POST:
                messages.success(request, 'Files and image notes saved successfully.')
                return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)

            # --- REDIRECT TO CHECKLIST INSTEAD OF AUTO FINALIZE ---
            messages.success(request, 'Files and image notes saved successfully. Please review and finalize.')
            return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)
    
    else:
        form = ConfirmationDocumentForm(instance=confirmation)
        image_formset = PackedImageFormSet(instance=confirmation)
    
    prev_url = reverse('clientdoc:edit_transport', kwargs={'invoice_id': invoice.id})
    packed_images_list = confirmation.packedimage_set.all()

    # Prepare available files for Checklist
    available_files = [
        {'id': 'invoice', 'name': 'Tax Invoice (Auto-Generated)', 'required': False},
    ]
    if hasattr(invoice, 'deliverychallan'):
        available_files.append({'id': 'dc', 'name': 'Delivery Challan (Auto-Generated)', 'required': False})
        
    if hasattr(invoice, 'transportcharges'):
        available_files.append({'id': 'transport', 'name': 'Transport Charges (Auto-Generated)', 'required': False})
    if has_po:
        available_files.append({'id': 'po', 'name': 'PO Copy (Uploaded)', 'required': False})
    if has_email:
        available_files.append({'id': 'email', 'name': 'Approval Email (Uploaded)', 'required': False})
    
    # Images are always last usually, but let's allow them in list if we want to be fancy, 
    # but for now images are appended at end in PDF gen logic typically. 
    # Let's keep images as a separate "Always at end" block or auto-included.
    
    context = {
        'form': form,
        'image_formset': image_formset,
        'invoice': invoice,
        'title': f'Confirmation & Finalize - Invoice {invoice.tally_invoice_number or invoice.id}',
        'has_po': has_po,
        'has_email': has_email,
        'prev_url': prev_url,
        'packed_images_list': packed_images_list, 
        'current_step': 4,
        'progress_percentage': 90, # Not 100 yet
        'available_files': available_files
    }
    return render(request, 'clientdoc/confirmation_checklist.html', context)


def finalize_invoice_pdf(request, invoice_id):
    """Generates the final PDF based on user selected order."""
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    confirmation = get_object_or_404(ConfirmationDocument, invoice=invoice)
    company_profile = OurCompanyProfile.objects.first()
    
    if request.method == 'POST':
        # Get order from POST
        # Valid separate IDs: invoice, dc, transport, po, email
        # We expect a comma separated string or list
        file_order_str = request.POST.get('file_order', 'invoice,dc,transport,po,email') 
        file_order = file_order_str.split(',')
        
        merger = PdfMerger()
        
        try:
            for file_type in file_order:
                if file_type == 'invoice':
                    invoice.calculate_total()
                    merger.append(generate_invoice_pdf(invoice, company_profile))
                
                elif file_type == 'dc' and hasattr(invoice, 'deliverychallan'):
                     merger.append(generate_dc_pdf(invoice, invoice.deliverychallan, company_profile))
                
                elif file_type == 'transport' and hasattr(invoice, 'transportcharges'):
                     merger.append(generate_transport_pdf(invoice, invoice.transportcharges, company_profile))
                
                elif file_type == 'po' and confirmation.po_file:
                    try:
                        PdfReader(confirmation.po_file.path)
                        merger.append(confirmation.po_file.path)
                    except Exception:
                        pass # Skip invalid
                
                elif file_type == 'email' and confirmation.approval_email_file:
                    try:
                        PdfReader(confirmation.approval_email_file.path)
                        merger.append(confirmation.approval_email_file.path)
                    except Exception:
                        pass

            # Always append images at the end ?? Or should they be in the order list?
            # User requirement: "all files... creating the PDF... selection... drag up and down".
            # Images are a collection of pages. Usually best at end.
            images_pdf_buffer = generate_packed_images_pdf(confirmation)
            if images_pdf_buffer:
                merger.append(images_pdf_buffer)
            
            output = BytesIO()
            merger.write(output)
            merger.close()
            output.seek(0)
            
            # Save logic ...
            filename_suffix = invoice.tally_invoice_number or invoice.id
            filename = f"confirmation_invoice_{filename_suffix}.pdf"
            
            # Preview vs Save
            path = os.path.join(settings.MEDIA_ROOT, 'confirmations', filename)
            os.makedirs(os.path.dirname(path), exist_ok=True) 

            with open(path, 'wb') as f:
                f.write(output.getvalue())
            
            confirmation.combined_pdf.name = f'confirmations/{filename}'
            confirmation.save()
            
            invoice.status = 'FIN'
            invoice.save()
            log_activity("Finalize Invoice", f"Finalized Invoice {invoice.tally_invoice_number or invoice.id}")
            
            messages.success(request, f'Document Bundle Generated Successfully!')
            return redirect('clientdoc:confirmation_list')

        except Exception as e:
            logger.error(f"Error finalizing PDF: {e}")
            messages.error(request, f"Error finalizing PDF: {e}")
            return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)
            
    return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)

# --- BULK UPLOAD VIEWS ---

def bulk_upload_page(request):
    """Page to upload excel and view history."""
    uploads = BulkInvoiceUpload.objects.order_by('-uploaded_at')
    
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        upload_type = request.POST.get('upload_type', 'invoice') # Default to invoice
        
        if not file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid Excel file.')
            return redirect('clientdoc:bulk_upload_page')
            
        upload_record = BulkInvoiceUpload.objects.create(file=file)
        upload_record.log = f"Type: {upload_type.title()}\n"
        upload_record.save()
        
        try:
            if upload_type == 'buyer':
                process_buyer_upload(upload_record)
            elif upload_type == 'item':
                process_item_upload(upload_record)
            elif upload_type == 'location':
                process_location_upload(upload_record)
            else:
                process_invoice_upload(upload_record)
                
            messages.success(request, f'{upload_type.title()} file uploaded and processed successfully.')
        except Exception as e:
            import traceback
            error_msg = f"Error processing file: {str(e)}\n{traceback.format_exc()}"
            upload_record.status = 'Failed'
            upload_record.log += error_msg
            upload_record.save()
            messages.error(request, 'Error processing file. Check logs.')
            
        return redirect('clientdoc:bulk_upload_page')
        
    return render(request, 'clientdoc/bulk_upload.html', {
        'uploads': uploads,
        'title': 'Bulk Data Upload'
    })

def download_sample_excel(request):
    """Generates a sample excel file based on type with formatting, optionally with data."""
    import datetime
    from openpyxl.styles import Font, PatternFill, Alignment
    
    upload_type = request.GET.get('type', 'invoice')
    do_export = request.GET.get('export') == 'true'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{upload_type.title()} {'Data' if do_export else 'Template'}"
    
    # Define Headers based on Type
    if upload_type == 'buyer':
        headers = ["Buyer Name*", "Address", "GSTIN", "State", "Phone", "Email"]
        widths = [30, 40, 20, 20, 20, 30]
        
    elif upload_type == 'item':
        headers = ["Item Name*", "Category", "Article/SKU", "Description", "Price*", "GST Rate (0.18)*", "HSN Code", "Unit (Nos)"]
        widths = [30, 20, 20, 40, 15, 15, 15, 15]
        
    elif upload_type == 'location':
        headers = ["Location Name*", "Site Code", "Address", "City", "State", "GSTIN", "Priority"]
        widths = [30, 15, 40, 20, 20, 20, 15]
        
    else: # Invoice
        headers = [
            "Buyer Name", "Location Name", "Item Name", "Quantity", "Generate Invoice (Yes/No)",
            "Generate PDF (Yes/No)", "App Invoice No. (For Update)",
            "Tally Invoice No.", 
            "Buyer's Order No.", "Buyer's Order Date (YYYY-MM-DD)", 
            "Dispatch Doc No.", "Dispatched Through", "Destination", 
            "Delivery Note", "Delivery Note Date (YYYY-MM-DD)", 
            "Mode/Terms of Payment", "Reference No. & Date", "Other References", 
            "Terms of Delivery", "Remarks",
            "DC Notes", "Transport Charges", "Transport Description"
        ]
        widths = [25, 25, 25, 10, 15, 15, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 25, 30, 30, 15, 30]

    ws.append(headers)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid") # Grey
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid") # Blue
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    # Blue First Column Header
    ws['A1'].fill = blue_fill
    
    # Set Widths
    for i, width in enumerate(widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # ---- EXPORT DATA LOGIC ----
    if do_export:
        if upload_type == 'buyer':
            for obj in Buyer.objects.all():
                ws.append([
                    obj.name, obj.address, obj.gstin, obj.state, obj.phone, obj.email
                ])
        elif upload_type == 'item':
            for obj in Item.objects.select_related('category').all():
                ws.append([
                    obj.name, 
                    obj.category.name if obj.category else "", 
                    obj.article_code, 
                    obj.description, 
                    obj.price, 
                    float(obj.gst_rate) if obj.gst_rate else 0.00,
                    obj.hsn_code, 
                    obj.unit
                ])
        elif upload_type == 'location':
            for obj in StoreLocation.objects.all():
                ws.append([
                    obj.name, obj.site_code, obj.address, obj.city, obj.state, obj.gstin, obj.priority
                ])
    
    # Invoice Specific Logic (Dropdowns etc - Only for Templates/Invoice)
    if upload_type == 'invoice':
        # Add Data and Validations
        data_ws = wb.create_sheet("Reference Data")
        data_ws.sheet_state = 'hidden' 
        
        buyers = list(Buyer.objects.values_list('name', flat=True))
        locations = list(StoreLocation.objects.values_list('name', flat=True))
        items = list(Item.objects.values_list('name', flat=True))
        
        for i, b in enumerate(buyers, 1): data_ws.cell(row=i, column=1, value=b)
        for i, l in enumerate(locations, 1): data_ws.cell(row=i, column=2, value=l)
        for i, it in enumerate(items, 1): data_ws.cell(row=i, column=3, value=it)

        def add_val(col, valid_range):
             dv = DataValidation(type="list", formula1=valid_range, allow_blank=True)
             ws.add_data_validation(dv)
             dv.add(f"{col}2:{col}500")

        if buyers: add_val('A', f"'Reference Data'!$A$1:$A${len(buyers)}")
        if locations: add_val('B', f"'Reference Data'!$B$1:$B${len(locations)}")
        if items: add_val('C', f"'Reference Data'!$C$1:$C${len(items)}")
        
        dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
        ws.add_data_validation(dv_yn)
        dv_yn.add("E2:E500")
        ws.add_data_validation(dv_yn) 
        dv_yn.add("F2:F500")
        
        # Defaults
        ws['E2'] = "Yes"
        ws['F2'] = "Yes" # Default PDF to Yes as requested
        ws['P2'] = "30 Days"
        ws['R2'] = "EMAIL Approval"
    
    # Timestamped Filename
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    mode = "Export" if do_export else "Template"
    filename = f"Bulk_{upload_type.title()}_{mode}_{timestamp}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

# --- PROCESSORS ---
def process_buyer_upload(record):
    ws = openpyxl.load_workbook(record.file.path, data_only=True).active
    log = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]: continue
        name = str(row[0]).strip()
        defaults = {
            'address': row[1] or "",
            'gstin': row[2] or "",
            'state': row[3] or "Karnataka",
            'phone': row[4] or "",
            'email': row[5] or ""
        }
        obj, created = Buyer.objects.update_or_create(name=name, defaults=defaults)
        log.append(f"Row {idx}: {'Created' if created else 'Updated'} Buyer '{name}'")
    
    record.log += "\n".join(log)
    record.status = 'Processed'
    record.save()

def process_item_upload(record):
    ws = openpyxl.load_workbook(record.file.path, data_only=True).active
    log = []
    from decimal import Decimal
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]: continue
        name = str(row[0]).strip()
        
        # Category Logic
        cat_name = row[1]
        category = None
        if cat_name:
            category, _ = ItemCategory.objects.get_or_create(name=str(cat_name).strip())
            
        price = 0.00
        try: price = float(row[4]) if row[4] else 0.00
        except: pass
        
        gst = 0.18
        try: gst = float(row[5]) if row[5] else 0.18
        except: pass

        defaults = {
            'category': category,
            'article_code': row[2] or "",
            'description': row[3] or "",
            'price': Decimal(price),
            'gst_rate': Decimal(gst),
            'hsn_code': row[6] or "844311",
            'unit': row[7] or "Nos"
        }
        obj, created = Item.objects.update_or_create(name=name, defaults=defaults)
        log.append(f"Row {idx}: {'Created' if created else 'Updated'} Item '{name}'")
        
    record.log += "\n".join(log)
    record.status = 'Processed'
    record.save()

def process_location_upload(record):
    ws = openpyxl.load_workbook(record.file.path, data_only=True).active
    log = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]: continue
        name = str(row[0]).strip()
        defaults = {
            'site_code': row[1] or "",
            'address': row[2] or "",
            'city': row[3] or "",
            'state': row[4] or "Karnataka",
            'gstin': row[5] or "",
            'priority': row[6] or ""
        }
        obj, created = StoreLocation.objects.update_or_create(name=name, defaults=defaults)
        log.append(f"Row {idx}: {'Created' if created else 'Updated'} Location '{name}'")
        
    record.log += "\n".join(log)
    record.status = 'Processed'
    record.save()

def process_invoice_upload(upload_record):
    """Parses Excel with PDF generation and Update support."""
    file_path = upload_record.file.path
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    log = []
    created_count = 0
    updated_count = 0
    error_count = 0
    
    from datetime import datetime
    
    def parse_date(date_val):
        if not date_val: return None
        if isinstance(date_val, datetime): return date_val
        try:
            return datetime.strptime(str(date_val).strip(), '%Y-%m-%d')
        except ValueError:
            return None 

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row): continue
        
        def get_col(idx):
            return row[idx] if idx < len(row) else None
            
        buyer_name = get_col(0)
        location_name = get_col(1)
        item_name = get_col(2)
        qty = get_col(3)
        gen_invoice = get_col(4)
        
        # New Cols
        gen_pdf = get_col(5)
        app_inv_no = get_col(6)
        
        # Header Fields (Shifted +2)
        tally_no = get_col(7)
        buyer_ord_no = get_col(8)
        buyer_ord_date = parse_date(get_col(9))
        disp_doc_no = get_col(10)
        disp_through = get_col(11)
        dest = get_col(12)
        del_note = get_col(13)
        del_note_date = parse_date(get_col(14))
        pay_terms = get_col(15) or "30 Days"
        ref_no = get_col(16)
        other_ref = get_col(17) or "EMAIL Approval"
        terms_del = get_col(18)
        remark = get_col(19)
        
        # DC & Transport
        dc_notes = get_col(20)
        trans_charges = get_col(21)
        trans_desc = get_col(22)

        # 1. Check Gen Invoice Flag
        if not gen_invoice or str(gen_invoice).strip().lower() != 'yes':
            log.append(f"Row {index}: Skipped (Generate = No)")
            continue
            
        # 2. Basic Validation
        if not (location_name and item_name and qty):
            log.append(f"Row {index}: Failed - Missing Location, Item or Qty")
            error_count += 1
            continue
            
        try:
            with transaction.atomic():
                # 3. Resolve FKs
                loc_obj = StoreLocation.objects.filter(name__iexact=str(location_name).strip()).first()
                if not loc_obj:
                    log.append(f"Row {index}: Failed - Location '{location_name}' not found")
                    error_count += 1
                    continue
                
                buyer_obj = None
                if buyer_name:
                    buyer_obj = Buyer.objects.filter(name__iexact=str(buyer_name).strip()).first()
                
                item_obj = Item.objects.filter(name__iexact=str(item_name).strip()).first()
                if not item_obj:
                     log.append(f"Row {index}: Failed - Item '{item_name}' not found")
                     error_count += 1
                     continue

                # 4. DETERMINE UPDATE OR CREATE
                invoice = None
                is_update = False
                
                if app_inv_no and str(app_inv_no).strip():
                    valid_id = str(app_inv_no).strip()
                    invoice = SalesInvoice.objects.filter(app_invoice_number__iexact=valid_id).first()
                    if invoice:
                        is_update = True
                
                # Header Data Dict
                header_data = {
                    'buyer': buyer_obj,
                    'location': loc_obj,
                    'tally_invoice_number': tally_no,
                    'buyers_order_no': buyer_ord_no,
                    'buyers_order_date': buyer_ord_date or datetime.now(),
                    'dispatch_doc_no': disp_doc_no,
                    'dispatched_through': disp_through,
                    'destination': dest,
                    'delivery_note': del_note,
                    'delivery_note_date': del_note_date or datetime.now(),
                    'mode_terms_payment': pay_terms,
                    'reference_no_date': ref_no,
                    'other_references': other_ref,
                    'terms_of_delivery': terms_del,
                    'remark': remark,
                }

                if is_update and invoice:
                    # Update fields
                    for key, val in header_data.items():
                        if val is not None: # Only update if value provided? Or overwrite? 
                            # Overwrite is safer for sync
                            setattr(invoice, key, val)
                    invoice.save()
                    log_entry_prefix = f"Row {index}: Updated Invoice {invoice.app_invoice_number}"
                    updated_count += 1
                else:
                    # Create New
                    header_data['date'] = datetime.now()
                    header_data['status'] = 'DRF'
                    invoice = SalesInvoice.objects.create(**header_data)
                    log_entry_prefix = f"Row {index}: Created Invoice #{invoice.id}"
                    created_count += 1
                
                # 5. Add Item (Append)
                InvoiceItem.objects.create(
                    invoice=invoice,
                    item=item_obj,
                    quantity=int(qty),
                    price=item_obj.price,
                )
                
                # 6. Auto-Create/Update DC
                if dc_notes:
                    dc, _ = DeliveryChallan.objects.get_or_create(invoice=invoice)
                    dc.notes = dc_notes
                    dc.save()
                    if invoice.status == 'DRF': invoice.status = 'DC'
                
                # 7. Auto-Create/Update Transport
                if trans_charges:
                    try:
                        amt = float(trans_charges)
                        trp, _ = TransportCharges.objects.get_or_create(invoice=invoice)
                        trp.charges = amt
                        trp.description = trans_desc
                        trp.save()
                        if invoice.status in ['DRF', 'DC']: invoice.status = 'TRP'
                    except ValueError:
                        log.append(f"{log_entry_prefix} - Warning: Invalid Transport Charge")

                invoice.save()
                invoice.calculate_total()
                
                # 8. GENERATE PDF?
                pdf_status = ""
                if gen_pdf and str(gen_pdf).strip().lower() == 'yes':
                    # Only gen if status allows (must have DC/TRP technically, but we can force)
                    try:
                        # Ensure confirmation doc exists
                        conf, _ = ConfirmationDocument.objects.get_or_create(invoice=invoice)
                        company_profile = OurCompanyProfile.objects.first()
                        
                        merger = PdfMerger()
                        
                        # 1. Invoice
                        invoice.calculate_total() 
                        merger.append(generate_invoice_pdf(invoice, company_profile))
                        
                        # 2. DC
                        if hasattr(invoice, 'deliverychallan'):
                            merger.append(generate_dc_pdf(invoice, invoice.deliverychallan, company_profile))
                            
                        # 3. Transport
                        if hasattr(invoice, 'transportcharges'):
                            merger.append(generate_transport_pdf(invoice, invoice.transportcharges, company_profile))
                            
                        # Output
                        output = BytesIO()
                        merger.write(output)
                        merger.close()
                        output.seek(0)
                        
                        # Save
                        filename_suffix = invoice.tally_invoice_number or invoice.app_invoice_number or str(invoice.id)
                        filename = f"confirmation_invoice_{filename_suffix}.pdf"
                        
                        # Django File Save
                        from django.core.files.base import ContentFile
                        conf.combined_pdf.save(filename, ContentFile(output.getvalue()), save=True)
                        
                        invoice.status = 'FIN'
                        invoice.save()
                        pdf_status = " + PDF Generated"
                        
                    except Exception as pdf_err:
                        pdf_status = f" + PDF Failed: {str(pdf_err)}"
                        logger.error(f"Bulk PDF Error: {pdf_err}")

                log.append(f"{log_entry_prefix}{pdf_status}")

        except Exception as e:
            log.append(f"Row {index}: Error - {str(e)}")
            error_count += 1
            
    upload_record.log = "\\n".join(log)
    upload_record.status = 'Processed'
    upload_record.save()

    
    return redirect('clientdoc:dashboard')

def create_buyer(request):
    if request.method == 'POST':
        form = BuyerForm(request.POST) 
        if form.is_valid():
            form.save()
            log_activity("Create Buyer", f"Created Buyer {form.instance.name}")
            messages.success(request, 'Buyer created successfully.')
            return redirect('clientdoc:dashboard')
    else:
        form = BuyerForm()
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Create Buyer'})

def buyer_list(request):
    search_fields = ['name', 'address', 'gstin', 'state']
    buyers = get_filtered_queryset(Buyer, request, search_fields)
    
    paginator = Paginator(buyers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/buyer_list.html', {
        'page_obj': page_obj, 
        'title': 'Buyer List',
        'list_type': 'buyer'
    })

def edit_buyer(request, pk):
    buyer = get_object_or_404(Buyer, pk=pk)
    if request.method == 'POST':
        form = BuyerForm(request.POST, instance=buyer)
        if form.is_valid():
            form.save()
            log_activity("Edit Buyer", f"Updated Buyer {buyer.name}")
            messages.success(request, 'Buyer updated successfully.')
            return redirect('clientdoc:buyer_list')
    else:
        form = BuyerForm(instance=buyer)
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Edit Buyer'})

def buyer_detail(request, pk):
    buyer = get_object_or_404(Buyer, pk=pk)
    return render(request, 'clientdoc/buyer_detail.html', {'buyer': buyer})


def delete_packed_image(request, image_id):
    """Handles the deletion of a specific packed image, ensuring file removal."""
    image = get_object_or_404(PackedImage, id=image_id)
    invoice_id = image.confirmation.invoice.id
    
    if request.method == 'POST':
        if image.image:
            image.image.delete(save=False) 
        
        image.delete()
        messages.success(request, 'Image successfully removed.')
    else:
        messages.error(request, 'Invalid request method.')
        
    return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)

def print_invoice(request, invoice_id):
    """Renders the print-friendly invoice template."""
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    company_profile = OurCompanyProfile.objects.first()
    
    # Ensure totals are calculated
    invoice.calculate_gst_totals()
    
    display_invoice_number = invoice.tally_invoice_number if invoice.tally_invoice_number else invoice.app_invoice_number
    
    # Determine IGST vs CGST/SGST based on model's calculated fields
    # Logic: If igst_total > 0, it's Inter-state. Or check place_of_supply vs company state.
    # However, model stores totals now.
    
    comp_state_code = company_profile.state_code if company_profile else '29'
    # Fallback to model POS if set, else Location state
    pos_code = invoice.place_of_supply if invoice.place_of_supply else (invoice.location.state_code if invoice.location else '29')
    
    is_igst = (pos_code != comp_state_code)
    
    # Re-sum taxable for display if needed, or rely on grand total - tax? 
    # Better to sum line items for the "Taxable Value" column/row in template.
    taxable_val = sum(item.taxable_value for item in invoice.invoiceitem_set.all())

    return render(request, 'clientdoc/invoice_print_template.html', {
        'invoice': invoice,
        'company': company_profile,
        'display_invoice_number': display_invoice_number,
        'taxable_val': taxable_val,
        'tax_amt': (invoice.cgst_total + invoice.sgst_total + invoice.igst_total),
        'cgst_amt': invoice.cgst_total,
        'sgst_amt': invoice.sgst_total,
        'igst_amt': invoice.igst_total,
        'is_igst': is_igst,
    })

def print_dc(request, invoice_id):
    """Renders the print-friendly Delivery Challan template."""
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    # Get the associated Delivery Challan
    dc = get_object_or_404(DeliveryChallan, invoice=invoice)
    company_profile = OurCompanyProfile.objects.first()
    
    # Calculate total quantity
    total_qty = sum(item.quantity for item in invoice.invoiceitem_set.all())
    display_invoice_number = invoice.tally_invoice_number if invoice.tally_invoice_number else invoice.app_invoice_number
    
    return render(request, 'clientdoc/dc_print_template.html', {
        'invoice': invoice,
        'dc': dc,
        'company': company_profile,
        'total_qty': total_qty,
        'display_invoice_number': display_invoice_number
    })

def print_transport(request, invoice_id):
    """Renders the print-friendly Transport Charges template."""
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    # Get the associated Transport Charges
    transport = get_object_or_404(TransportCharges, invoice=invoice)
    company_profile = OurCompanyProfile.objects.first()
    
    display_invoice_number = invoice.tally_invoice_number if invoice.tally_invoice_number else invoice.app_invoice_number
    
    return render(request, 'clientdoc/transport_print_template.html', {
        'invoice': invoice,
        'transport': transport,
        'company': company_profile,
        'display_invoice_number': display_invoice_number
    })
def project_guide(request):
    """Serves the Project Guide PDF."""
    import os
    from django.conf import settings
    from django.http import HttpResponse, Http404

    file_path = os.path.join(settings.BASE_DIR, 'Project guide', 'Project Guide.pdf')
    if os.path.exists(file_path):
        with open(file_path, 'rb') as pdf:
            response = HttpResponse(pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="Project Guide.pdf"'
            return response
    else:
        raise Http404("Project Guide not found")
